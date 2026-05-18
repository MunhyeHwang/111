import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# =========================
# 1. 读取
# =========================
def read_annotation(name):
    csv_path = f'sentiment_annotation_标注员{name}.csv'
    xlsx_path = f'sentiment_annotation_标注员{name}.xlsx'
    if os.path.exists(csv_path):
        return pd.read_csv(csv_path, encoding='utf-8-sig')
    elif os.path.exists(xlsx_path):
        return pd.read_excel(xlsx_path)
    else:
        raise FileNotFoundError(f'找不到 {csv_path}')

df_A = read_annotation('A')
df_B = read_annotation('B')
df_C = read_annotation('C')

# =========================
# 2. 清理标签为数值
# =========================
for df in [df_A, df_B, df_C]:
    df['情感标注'] = pd.to_numeric(df['情感标注'], errors='coerce')
    df['评论'] = df['评论'].astype(str).str.strip()

# =========================
# 3. 提取共标(用评论文本作为主键)
# =========================
shared_A = df_A[df_A['任务类型']=='共标'][['评论', '情感标注']].rename(columns={'情感标注': 'A_情感'})
shared_B = df_B[df_B['任务类型']=='共标'][['评论', '情感标注']].rename(columns={'情感标注': 'B_情感'})
shared_C = df_C[df_C['任务类型']=='共标'][['评论', '情感标注']].rename(columns={'情感标注': 'C_情感'})

# 去重(同一标注员里如果有重复评论,保留第一个)
shared_A = shared_A.drop_duplicates(subset=['评论'], keep='first')
shared_B = shared_B.drop_duplicates(subset=['评论'], keep='first')
shared_C = shared_C.drop_duplicates(subset=['评论'], keep='first')

# 用评论文本 inner join 三人标注
shared_merged = shared_A.merge(shared_B, on='评论', how='inner').merge(shared_C, on='评论', how='inner')
print(f'三人共同标注的评论数(用文本对齐): {len(shared_merged)}')

# 剔除任意一人未标或标错的样本
shared_merged = shared_merged.dropna(subset=['A_情感', 'B_情感', 'C_情感'])
shared_merged = shared_merged[
    shared_merged['A_情感'].isin([0, 1]) &
    shared_merged['B_情感'].isin([0, 1]) &
    shared_merged['C_情感'].isin([0, 1])
].copy()
shared_merged[['A_情感', 'B_情感', 'C_情感']] = shared_merged[['A_情感', 'B_情感', 'C_情感']].astype(int)
print(f'有效共标(三人都正确标注): {len(shared_merged)}')

# =========================
# 4. Fleiss Kappa 计算
# =========================
n_items = len(shared_merged)
n_categories = 2

matrix = np.zeros((n_items, n_categories), dtype=int)
for i, row in enumerate(shared_merged.itertuples(index=False)):
    labels = [row.A_情感, row.B_情感, row.C_情感]
    matrix[i, 0] = labels.count(0)
    matrix[i, 1] = labels.count(1)

def fleiss_kappa(matrix):
    N, k = matrix.shape
    n = matrix.sum(axis=1)[0]
    P_i = (np.sum(matrix ** 2, axis=1) - n) / (n * (n - 1))
    P_bar = np.mean(P_i)
    p_j = matrix.sum(axis=0) / (N * n)
    P_e = np.sum(p_j ** 2)
    if P_e == 1:
        return 1.0
    return (P_bar - P_e) / (1 - P_e)

def interpret_kappa(k):
    if k < 0: return '差于随机一致 (Poor)'
    elif k < 0.20: return '轻微一致 (Slight)'
    elif k < 0.40: return '一般一致 (Fair)'
    elif k < 0.60: return '中等一致 (Moderate)'
    elif k < 0.80: return '高度一致 (Substantial) ✓ 达标'
    else: return '几乎完全一致 (Almost Perfect) ✓✓ 优秀'

kappa = fleiss_kappa(matrix)
print(f'\nFleiss Kappa = {kappa:.4f}')
print(f'解释: {interpret_kappa(kappa)}')

# =========================
# 5. 共标最终情感(多数投票)
# =========================
def decide(row):
    labels = [row['A_情感'], row['B_情感'], row['C_情感']]
    c1, c0 = labels.count(1), labels.count(0)
    if c1 == 3: return 1, '完全一致(好评)'
    elif c0 == 3: return 0, '完全一致(差评)'
    elif c1 == 2: return 1, '2人一致(好评)'
    else: return 0, '2人一致(差评)'

results = shared_merged.apply(decide, axis=1, result_type='expand')
shared_merged['最终情感'] = results[0]
shared_merged['一致性'] = results[1]

print(f'\n=== 共标一致性分布 ===')
print(shared_merged['一致性'].value_counts())

# =========================
# 6. 处理"非共标"部分(不去重版)
# =========================
# 三人各自所有标注
all_A = df_A[['评论', '情感标注']].rename(columns={'情感标注': '最终情感'}).copy()
all_A['来源'] = 'A'
all_B = df_B[['评论', '情感标注']].rename(columns={'情感标注': '最终情感'}).copy()
all_B['来源'] = 'B'
all_C = df_C[['评论', '情感标注']].rename(columns={'情感标注': '最终情感'}).copy()
all_C['来源'] = 'C'

all_annotations = pd.concat([all_A, all_B, all_C], ignore_index=True)
all_annotations = all_annotations.dropna(subset=['最终情感'])
all_annotations['最终情感'] = all_annotations['最终情感'].astype(int)
all_annotations = all_annotations[all_annotations['最终情感'].isin([0, 1])]

# 共标 422 条用多数投票结果替换原始三人标注
# 策略:
# - 凡是出现在共标集合中的评论,从 all_annotations 中剔除三人的原始标注
# - 用多数投票后的单一标注替代
shared_decisions = shared_merged[['评论', '最终情感']].copy()
shared_decisions['来源'] = '共标(多数投票)'
shared_comments_set = set(shared_decisions['评论'].tolist())

# 不在共标集合的部分:保留所有重复标注
non_shared = all_annotations[~all_annotations['评论'].isin(shared_comments_set)].copy()

# 拼接:共标(多数投票后的) + 非共标(保留所有重复标注)
final_dataset = pd.concat([shared_decisions, non_shared], ignore_index=True)

print(f'\n=== 最终标注数据集(保留重复评论) ===')
print(f'总条数: {len(final_dataset)}')
print(f'好评: {(final_dataset["最终情感"]==1).sum()} 条 ({(final_dataset["最终情感"]==1).mean()*100:.2f}%)')
print(f'差评: {(final_dataset["最终情感"]==0).sum()} 条 ({(final_dataset["最终情感"]==0).mean()*100:.2f}%)')

# 保存
final_dataset[['评论', '最终情感']].to_csv('sentiment_annotated_final.csv', index=False, encoding='utf-8-sig')
print(f'\n已保存: sentiment_annotated_final.csv')

# =========================
# 7. Excel 报告
# =========================
wb = Workbook()
ws = wb.active
ws.title = 'Kappa检验报告'

report = [
    ['指标', '数值/说明'],
    ['标注员人数', 3],
    ['每人总标注条数', 1500],
    ['', ''],
    ['计算 Kappa 的有效共标条数', len(shared_merged)],
    ['Fleiss Kappa', round(kappa, 4)],
    ['Kappa 解释', interpret_kappa(kappa)],
    ['', ''],
    ['完全一致条数', sum('完全一致' in x for x in shared_merged['一致性'])],
    ['2人一致条数', sum('2人一致' in x for x in shared_merged['一致性'])],
    ['', ''],
    ['最终数据集总条数', len(final_dataset)],
    ['好评条数', (final_dataset['最终情感']==1).sum()],
    ['差评条数', (final_dataset['最终情感']==0).sum()],
]

for row_idx, row_data in enumerate(report, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
wb.save('sentiment_kappa_report.xlsx')
print('已保存: sentiment_kappa_report.xlsx')